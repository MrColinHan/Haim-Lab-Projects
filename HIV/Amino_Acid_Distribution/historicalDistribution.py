#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Jun 25 13:42:51 2018

@author: Changze Han
"""


import xlrd
import openpyxl

'''
    !!!!Remember to refactor this script in order to improve efficiency!!!
'''
'''
This program calculates the historical distribution of different positions' amino acids
  STEPS:
1. Change the Input directory, do not delete r 
2. Change the Output directory, do not delete r, When you copy the address, keep in mind there's also a slash at the end.
3. Make a name of the Output file
4. Change the PositionRange: if you want to select a single position: [[88,88]]
                             if you want to select multiple positions: [[88,88],[156,156],[276,276]]
                             if you want to select a position range, like from 88 to 637: [[88,637]]
                             you can also mix them: [[88,88],[156,156],[197,295],[392,637]]
5. Change the YearRange, the format is the same as the PositionRange:
                                 single year:[1979,1979]
                                 year range: [[1979,2014]]
                                 mix:[[1979,2014],[1985,2006],[1994,2000],[2008,2008]]
6. Ignore the output file named 'ignore.xlsx', feel free to delete it
7. run Forest run!
'''

# input format : Country Year Patient Accession 131 132 133 ......
Input = r"/Users/Han/Documents/Haim_Lab(2018_summer)/raw_data/gp120_pngs/7.12.19_B_GP120_PNGS.xlsx"#r"/Users/Han/Documents/Haim_Lab(2018_summer)/9.27.20_FPPR-FD_Covid-updatedcbElisa/(Combined-AminoAcid)Clade C AFR FPPR.xlsx"
Output = r"/Users/Han/Documents/Haim_Lab(2018_summer)/5.10.21 B gp120 dynaFD/"
OutputName ="Clade B gp120 Dyna-FD.xlsx"
TabNumber = 0 # tab number in Excel file, start as index 0
PositionRange = [[131, 511]]
#pngs:[[88,88],[156,156],[160,160],[197,197],[230,230],[234,234],[241,141],[262,262],
#[276,276],[289,289],[295,295],[301,301],[332,332],[339,339],[356,356],[362,362],[386,386],
#[392,392],[448,448],[133,133],[137,137],[185,185],[355,355],[398,398],[406,406],[411,411],[462,462]]
#2g12: [[295,295],[332,332],[339,339],[392,392],[448,448]]
#2F5: [[662,662],[663,663],[664,664],[665,665],[667,667]]
YearRange = [[1979,1986],[1987,1994],[1995,1999],[2000,2004],[2005,2009],[2010,2015]]
## CURRENT: [[2007,2015]]
# B_NA: 300
# B_EU: 164
# B_AS-kr: [[2005,2015]]   150
# B_KR: [[2001,2009]]   293
# C_ECA: 103
# C_SA: 239
# C_EU: [[2005,2015]]   100
# C_INNP: [[2000,2015]] 68
# AE_CN: 156
# AE_TH: 114
# A1_all: [[2006,2015]]   129
##DYNAMICS: 
# Clade B: [[1979,1986],[1987,1994],[1995,1999],[2000,2004],[2005,2009],[2010,2015]] 
# Clade C: [[1986,1994],[1995,1999],[2000,2004],[2005,2009],[2010,2015]] 
# Clade AE:[[1986,1994],[1995,1999],[2000,2004],[2005,2007],[2008,2015]] 
# Clade A1:[[1986,1994],[1995,1999],[2000,2004],[2005,2009],[2010,2015]]  
# B_Korea: [[1991,1996],[1997,2000],[2001,2004],[2005,2009]] 
####################################################################################################################################
####################################################################################################################################
####################################################################################################################################
####################################################################################################################################
####################################################################################################################################
#Code:

workbook = xlrd.open_workbook(Input)
sheet = workbook.sheet_by_index(TabNumber)
nRows = sheet.nrows
nCols = sheet.ncols


#get data of row y by index, eg:GetRowIndexData(2)
def GetRowIndexData(y):
    row = []
    for col in range (nCols):
        row.append(sheet.cell_value((y-1),col))
    return row

#get data of column x by index
def GetColIndexData(x):
    col = []
    for row in range (nRows):
        col.append(sheet.cell_value(row,(x-1)))
    return col

#get data of column x by title of first row , eg: GetColData('Year')
def GetColData(x): 
    
    for i in GetRowIndexData(1):
        if i == x : 
            a = GetColIndexData(GetRowIndexData(1).index(i)+1)
            del a[0]
            return a
#get data at a specific location, eg:GetDataAt(1,2)
def GetDataAt(y,x):
    return (sheet.cell_value((y-1),(x-1)))


#get position List
def getPositionList():
    result = []
    for i in PositionRange:
        for j in GetRowIndexData(1):
            if (type(j)== int) or (type(j)== float):
                if (j>= i[0]) and (j <= i[1]) :
                    result.append(int(j))
    return result
            
    
Position = getPositionList()

#get data of position p during year range y
def GetYearRangeData(y,p):
    finalfinal=[]
    for k in p:
        pos = GetRowIndexData(1).index(k)+1    
        final = []
        yearRange = GetColData('Year')
        for i in y:
            result = []
            count = 1
            for j in yearRange: 
                count = count + 1
                if (j >= i[0]) and (j <= i[1]) :
                    result.append(GetDataAt(count,pos))
                    
            final.append(result)
        finalfinal.append(str(k))
        finalfinal.append(final)
        
        
    return finalfinal
   
     

test2 = GetYearRangeData(YearRange,Position)                                         


#write into excel file

def writeFile():
    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()
    sheet.title='Sheet #1'

#Loop to set the value of each cell
    rCount = 1
    cCount = 1
    for i in range(len(test2)):
        #print (i)
        if type(test2[i]) == str:
            sheet.cell(row=rCount, column=cCount).value = test2[i]
            cCount = cCount + 1
        else:
            for j in range(len(test2[i])):
            #print(j)
                for k in range(len(test2[i][j])):
                
                    sheet.cell(row=rCount, column=cCount).value=test2[i][j][k]
                    rCount = rCount + 1
                rCount = 1
                cCount = cCount + 1
    wb.save(Output+"Ignore.xlsx")
    
writeFile()


#open the ignore.xlsx, then count the amino acid
file_location2 = Output+"ignore.xlsx"
workbook2 = xlrd.open_workbook(file_location2)
sheet2 = workbook2.sheet_by_index(0)
nRows2 = sheet2.nrows
nCols2 = sheet2.ncols

def GetColIndexData2(x):
    col = []
    for row in range (nRows2):
        col.append(sheet2.cell_value(row,(x-1)))
    return col

def Count():
    #name = ['A','C','D','E','F','G','H','I','K','L','M','N','P','Q','R','S','T','V','W','Y','Z']
    name = ['Z','N','T','S','D','E','K','R','H','Y','Q','I','L','V','A','C','F','G','M','P','W']
    final = []
    for col in range(nCols2):
        colCount = []
        if GetColIndexData2(col+1)[0].isdigit() :
            final.append(GetColIndexData2(col+1)[0])
            final.append(name)
        else:
            for aa in name:      
                num = 0
                for n in GetColIndexData2(col+1):
                
                    if aa == n:
                        num = num + 1
                colCount.append(num)
            final.append(colCount)                           
    return final

test3 = Count()
test4 = Count()
# add one row which is the total number of amino acids of each position 
def sum():  
    for i in range(len(test4)):       
        count = 0
        if type(test4[i]) is str:
            test4[i] = ''
        elif type(test4[i][0]) is str:
                    test4[i] = "Sum"
        else:
            for j in test4[i]:
                count = count + j
            test4[i] = count
    return test4

test4 = sum()

#write into excel file
def writeFile2():
    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()
    sheet.title='Sheet #1'

#Loop to set the value of each cell
    colC = 3
    for p in range(len(Position)):
        
        for i in range(len(YearRange)):
            sheet.cell(row=1, column=colC).value=str(YearRange[i])
            colC = colC+1
        colC = colC + 2
    colCC = 1
    for i in range(len(test3)):
        
        if type(test3[i]) == str:  
            colCC = colCC+1
            sheet.cell(row=1, column=colCC).value=test3[i]            
        else:
            for j in range(len(test3[i])):
                sheet.cell(row=j+2, column=colCC).value=test3[i][j]
            colCC = colCC+1
    for i in range(len(test4)):
        sheet.cell(row=24, column=i+1).value=test4[i]
    wb.save(Output + OutputName)
    
writeFile2()






